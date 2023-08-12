import openpyxl
import time
from selenium import webdriver
from datetime import datetime as date

driver = webdriver.ChromiumEdge(executable_path=r"D:\Selenium\webdriver\msedgedriver.exe")

def scrapfunc(e):
    global long, short
    driver.maximize_window()
    driver.get("https://www.google.com/")
    sbox = driver.find_element("xpath", ".//*[@title='সার্চ করুন']")
    sbox.send_keys(e)
    time.sleep(5)
    sug = driver.find_elements("xpath","//div[@class='wM6W7d' and @role='presentation']/span")
    for row in sug:
        x.append(row.text)
    xx = list(filter(None, x))
    long = max(xx, default=None)
    short = min(xx, default=None)
    x.clear()
    xx.clear()
    return long, short

day = date.today().strftime("%A")
kw = []
x = []

path = r"D:\Selenium\excel\excel.xlsx"
wb = openpyxl.load_workbook(path)
ws = wb[day]

for cell in ws['C']:
    if cell.value is not None:
        kw.append(cell.value)

i = 0
while i < len(kw):
    scrapfunc(kw[i])
    cv = str(3 + i)
    ws['D' + cv] = long
    ws['E' + cv] = short
    wb.save(r'D:\Selenium\excel\excel.xlsx')
    i += 1
driver.quit()